import socket
from datetime import datetime
from io import BytesIO

from django.db.models import Count, Sum
from django.http import HttpResponse, JsonResponse
from django.shortcuts import render
from django.views.decorators.csrf import ensure_csrf_cookie
from django.utils import timezone
from openpyxl import Workbook
from openpyxl.styles import Font
from rest_framework.decorators import api_view, permission_classes
from rest_framework.permissions import AllowAny

from academics.models import Unit
from attendance.models import AttendanceSession
from institutions.models import Institution
from trainers.models import Trainer


def get_lan_addresses(request):
    port = request.get_port() or "8000"
    addresses = []
    seen = set()

    def add_address(ip_address):
        if not ip_address or ip_address.startswith("127."):
            return
        if ip_address in seen:
            return
        seen.add(ip_address)
        addresses.append(
            {
                "ip": ip_address,
                "backend_url": f"http://{ip_address}:{port}",
                "health_url": f"http://{ip_address}:{port}/api/health/",
            }
        )

    hostname = socket.gethostname()
    try:
        for result in socket.getaddrinfo(hostname, None, socket.AF_INET):
            add_address(result[4][0])
    except socket.gaierror:
        pass

    try:
        with socket.socket(socket.AF_INET, socket.SOCK_DGRAM) as sock:
            sock.connect(("8.8.8.8", 80))
            add_address(sock.getsockname()[0])
    except OSError:
        pass

    return {
        "hostname": hostname,
        "request_host": request.get_host(),
        "current_request_url": request.build_absolute_uri("/"),
        "addresses": addresses,
    }


@api_view(["GET"])
@permission_classes([AllowAny])
def health_check(request):
    return JsonResponse({"status": "ok", "service": "clock-in-out-backend"})


@api_view(["GET"])
@permission_classes([AllowAny])
def network_info(request):
    return JsonResponse(get_lan_addresses(request))


@ensure_csrf_cookie
def monitor_dashboard(request):
    active_sessions = AttendanceSession.objects.filter(is_active=True).select_related(
        "trainer",
        "unit",
    )[:8]
    context = {
        "network": get_lan_addresses(request),
        "institution_count": Institution.objects.count(),
        "trainer_count": Trainer.objects.count(),
        "unit_count": Unit.objects.count(),
        "active_session_count": AttendanceSession.objects.filter(is_active=True).count(),
        "recent_sessions": AttendanceSession.objects.select_related("trainer", "unit")[:8],
        "active_sessions": active_sessions,
    }
    return render(request, "monitor.html", context)


@ensure_csrf_cookie
def landing_page(request):
    return render(request, "landing.html")


@ensure_csrf_cookie
def web_dashboard(request):
    context = {
        "institution_count": Institution.objects.count(),
        "trainer_count": Trainer.objects.count(),
        "unit_count": Unit.objects.count(),
        "active_session_count": AttendanceSession.objects.filter(is_active=True).count(),
        "recent_sessions": AttendanceSession.objects.select_related("trainer", "unit")[:8],
    }
    return render(request, "web/dashboard.html", context)


@ensure_csrf_cookie
def setup_page(request):
    return render(request, "web/setup.html")


@ensure_csrf_cookie
def clock_page(request):
    return render(request, "trainer/clock.html")


@ensure_csrf_cookie
def trainers_page(request):
    return render(request, "web/trainers.html")


@ensure_csrf_cookie
def units_page(request):
    return render(request, "web/units.html")


@ensure_csrf_cookie
def assignments_page(request):
    return render(request, "web/assignments.html")


@ensure_csrf_cookie
def active_classes_page(request):
    active_sessions = AttendanceSession.objects.filter(is_active=True).select_related(
        "trainer",
        "unit",
    )
    return render(request, "web/active_classes.html", {"active_sessions": active_sessions})


@ensure_csrf_cookie
def attendance_history_page(request):
    sessions = AttendanceSession.objects.select_related("trainer", "unit")[:100]
    return render(request, "web/attendance_history.html", {"sessions": sessions})


def report_sessions_queryset(request):
    queryset = AttendanceSession.objects.select_related("trainer", "unit", "institution")
    start_date = request.GET.get("start_date")
    end_date = request.GET.get("end_date")
    trainer_id = request.GET.get("trainer")
    unit_id = request.GET.get("unit")

    if start_date:
        start = timezone.make_aware(datetime.strptime(start_date, "%Y-%m-%d"))
        queryset = queryset.filter(clock_in_at__gte=start)
    if end_date:
        end = timezone.make_aware(datetime.strptime(end_date, "%Y-%m-%d")).replace(
            hour=23,
            minute=59,
            second=59,
        )
        queryset = queryset.filter(clock_in_at__lte=end)
    if trainer_id:
        queryset = queryset.filter(trainer_id=trainer_id)
    if unit_id:
        queryset = queryset.filter(unit_id=unit_id)

    return queryset


@ensure_csrf_cookie
def reports_page(request):
    sessions = report_sessions_queryset(request)
    trainer_summary = (
        sessions.filter(is_active=False)
        .values("trainer__id", "trainer__name", "trainer__id_number")
        .annotate(
            session_count=Count("id"),
            actual_minutes_total=Sum("actual_minutes"),
            credited_minutes_total=Sum("credited_minutes"),
        )
        .order_by("trainer__name")
    )
    unit_summary = (
        sessions.filter(is_active=False)
        .values("unit__id", "unit__code", "unit__name")
        .annotate(
            actual_minutes_total=Sum("actual_minutes"),
            credited_minutes_total=Sum("credited_minutes"),
        )
        .order_by("unit__code")
    )
    context = {
        "trainers": Trainer.objects.order_by("name"),
        "units": Unit.objects.order_by("code"),
        "sessions": sessions[:100],
        "trainer_summary": trainer_summary,
        "unit_summary": unit_summary,
        "filters": request.GET,
    }
    return render(request, "web/reports.html", context)


def append_headers(sheet, headers):
    sheet.append(headers)
    for cell in sheet[1]:
        cell.font = Font(bold=True)


def minutes_to_hours(minutes):
    return round((minutes or 0) / 60, 2)


def build_attendance_workbook(request):
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Attendance"
    append_headers(
        sheet,
        [
            "Trainer",
            "ID Number",
            "Unit Code",
            "Unit Name",
            "Clock In",
            "Clock Out",
            "Actual Minutes",
            "Credited Minutes",
            "Actual Hours",
            "Credited Hours",
            "Status",
        ],
    )
    for session in report_sessions_queryset(request):
        sheet.append(
            [
                session.trainer.name,
                session.trainer.id_number,
                session.unit.code,
                session.unit.name,
                session.clock_in_at,
                session.clock_out_at,
                session.actual_minutes,
                session.credited_minutes,
                minutes_to_hours(session.actual_minutes),
                minutes_to_hours(session.credited_minutes),
                "Active" if session.is_active else "Closed",
            ]
        )
    return workbook


def build_trainer_summary_workbook(request):
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Trainer Summary"
    append_headers(
        sheet,
        [
            "Trainer",
            "ID Number",
            "Sessions",
            "Actual Minutes",
            "Credited Minutes",
            "Actual Hours",
            "Credited Hours",
        ],
    )
    rows = (
        report_sessions_queryset(request)
        .filter(is_active=False)
        .values("trainer__name", "trainer__id_number")
        .annotate(
            session_count=Count("id"),
            actual_minutes_total=Sum("actual_minutes"),
            credited_minutes_total=Sum("credited_minutes"),
        )
        .order_by("trainer__name")
    )
    for row in rows:
        actual = row["actual_minutes_total"] or 0
        credited = row["credited_minutes_total"] or 0
        sheet.append(
            [
                row["trainer__name"],
                row["trainer__id_number"],
                row["session_count"],
                actual,
                credited,
                minutes_to_hours(actual),
                minutes_to_hours(credited),
            ]
        )
    return workbook


def build_unit_summary_workbook(request):
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Unit Summary"
    append_headers(
        sheet,
        [
            "Unit Code",
            "Unit Name",
            "Actual Minutes",
            "Credited Minutes",
            "Actual Hours",
            "Credited Hours",
        ],
    )
    rows = (
        report_sessions_queryset(request)
        .filter(is_active=False)
        .values("unit__code", "unit__name")
        .annotate(
            actual_minutes_total=Sum("actual_minutes"),
            credited_minutes_total=Sum("credited_minutes"),
        )
        .order_by("unit__code")
    )
    for row in rows:
        actual = row["actual_minutes_total"] or 0
        credited = row["credited_minutes_total"] or 0
        sheet.append(
            [
                row["unit__code"],
                row["unit__name"],
                actual,
                credited,
                minutes_to_hours(actual),
                minutes_to_hours(credited),
            ]
        )
    return workbook


def excel_response(workbook, filename):
    output = BytesIO()
    workbook.save(output)
    output.seek(0)
    response = HttpResponse(
        output.getvalue(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    response["Content-Disposition"] = f'attachment; filename="{filename}"'
    return response


def export_attendance_report(request):
    return excel_response(build_attendance_workbook(request), "attendance_report.xlsx")


def export_trainer_summary(request):
    return excel_response(build_trainer_summary_workbook(request), "trainer_summary.xlsx")


def export_unit_summary(request):
    return excel_response(build_unit_summary_workbook(request), "unit_summary.xlsx")
